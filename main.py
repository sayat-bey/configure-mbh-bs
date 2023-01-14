import os
import yaml
from sys import argv
from datetime import datetime
from pathlib import Path
from netmiko import ConnectHandler
from jinja2 import Environment, FileSystemLoader
from openpyxl import load_workbook

#######################################################################################
# ------------------------------                   -----------------------------------#
#######################################################################################

helper = {
    "kyzy": ["172.20.50.22", "172.20.24.170"],
    "alma": ["172.20.17.181", "172.20.17.209", "172.20.24.170", "172.20.0.178"],
    "shim": ["172.20.18.117", "172.20.24.170"],
    "tara": ["172.20.22.157", "172.20.22.161", "172.20.24.170"],
    "seme": ["172.20.14.33", "172.20.24.170"],
    "ural": ["172.20.12.33", "172.20.24.170"],
    "akta": ["172.20.15.33", "172.20.24.170"],
    "kost": ["172.20.11.33", "172.20.24.170"],
    "asta": ["172.20.26.14", "172.20.19.9", "172.20.24.170"],
    "koks": ["172.20.9.33", "172.20.24.170"],
    "petr": ["172.20.10.33", "172.20.24.170"],
    "pavl": ["172.20.36.54", "172.20.24.170"],
    "ustk": ["172.20.46.37", "172.20.24.170"],
    "kara": ["172.20.34.2", "172.20.24.170"],
    "akto": ["172.20.28.2", "172.20.24.170"],
    "atyr": ["172.20.30.38", "172.20.24.170"]
    }

conf_dic = {
    "helpers": None,    # ["172.20.14.33", "172.20.24.170"]
    "vlans": None,      # [1010 1011 1012 .. 1015]
    "port": None,       # BS port
    "ip": None,         # [list of bs ip]
    "mask": None,       # mask of bs
    "ios_type": None,   # cisco_ios, cisco_xr, cisco_xe
    "errors": [],       # errors
    }

configuration_log = []


#######################################################################################
# ------------------------------ def function part -----------------------------------#
#######################################################################################


def get_arguments(arguments):
    cfg = False
    for arg in arguments:
        if arg == "cfg" or arg == "CFG" or arg == "conf":
            cfg = True
    print(f"config mode:...................{cfg}")
  
    return cfg


def get_user_pw():
    with open("psw.yaml") as file:
        user_psw = yaml.load(file, yaml.SafeLoader)

    return user_psw[0], user_psw[1]


def set_ios(conf_dic, device):
    while True:
        if "csg" in device:
            ios_type = input("Enter ios type (ios,xr,xe) [ios]: ") or "ios"
        elif "pagg" in device:  
            ios_type = input("Enter ios type (ios,xr,xe) [xr]: ") or "xr"
        else:
            ios_type = input("Enter ios type (ios,xr,xe) [ios]: ") or "ios"

        if ios_type == "ios":
            conf_dic["ios_type"] = "cisco_ios"
            break
        elif ios_type == "xr":
            conf_dic["ios_type"] = "cisco_xr"
            break
        elif ios_type == "xe":
            conf_dic["ios_type"] = "cisco_xe"
            break
        else:
            print(f"# ERROR # Wrong ios type: {ios_type}. It must be one of ios,xr,xe")


def connect(usr, psw, conf_dic, device):
    ssh_conn = ConnectHandler(device_type=conf_dic["ios_type"], ip=device, username=usr, password=psw)
    show_inf_desc = ssh_conn.send_command("show interfaces description")

    return ssh_conn, show_inf_desc


def define_hostname(ssh_conn, device):
    if len(device.split(".")) == 4:    # device in ip 
        prmt = ssh_conn.find_prompt()
        hostname = prmt.split("#")[0]
        print(f"hostname is: {hostname}")
    else:
        hostname = device

    region = hostname.split("-")[0]
    if "." in region:
        region_final = region.split(".")[1]
    else:
        region_final = region    

    return region_final, hostname


def set_region(conf_dic, probable_region, helper):
    while True:
        region = input(f"Enter region for IP RELAY [{probable_region}]: ") or probable_region
        if helper.get(region):
            break
        else:
            print("wrong region. enter one of:\n\
                kyzy, alma, shim, tara, seme, ural, akta, kost\n\
                asta, koks, petr, pavl, ustk, kara, akto, atyr")

    conf_dic["helpers"] = helper[region]


def define_port_vlan(show_inf_desc, conf_dic):
    print(f"{show_inf_desc}")
    print("----------------------------------------------------------------------")
    vl = []     # 101 102 103
    probable_last_vlan = "1010"

    if conf_dic["ios_type"] == "cisco_ios":
        for line in show_inf_desc.splitlines():
            if "Vl10" in line:
                vl.append(line.split()[0][2:5])     #Vl1010 -> 101

    elif conf_dic["ios_type"] == "cisco_xr":  
        for line in show_inf_desc.splitlines():
            if ".10" in line:
                vl.append(line.split()[0][-4:-1])     #interface.1010 -> 101

    if len(vl) > 1:
        last_vlan = vl[-1]
        probable_last_vlan = str(int(last_vlan)+1) + "0"

    vlan = input(f"Enter first vlan (1010,1020,..,10x0) [{probable_last_vlan}]: ") or probable_last_vlan
    vlans = [str(int(vlan)+i) for i in range(6)]
    conf_dic["vlans"] = vlans
    conf_dic["port"] = input("Enter BS interface: ")
    

def load_excel(conf_dic):
    cwd = os.getcwd()
    files = []
    ip_list = []
    x = 1
    y = 1

    for i in os.listdir(cwd):
        if os.path.isfile(i) and "xlsx" in i:
            files.append(i)

    print(f"xlsx files in current directory: [{', '.join(files)}]")	
    probable_file = files[0]
    excel_file = input(f"Enter BS excel file [{probable_file}]: ") or probable_file
    wb = load_workbook(excel_file)
    first_sheet = wb.sheetnames[0]
    sheet = wb[first_sheet]

    while True:
        v = sheet.cell(row=x, column=y).value
        if "." in str(v) and len(v.split(".")) == 4:
            last_octet1 = v.split(".")[3]
            v2 = sheet.cell(row=x, column=y+1).value
            v3 = sheet.cell(row=x, column=y+2).value
            v4 = sheet.cell(row=x, column=y+3).value

            # v - ip csg, v2- ip bs, v3 - mask
            if "." in str(v2) and "255.255.255." in str(v3):
                last_octet2 = v2.split(".")[3]
                if int(last_octet1) + 1 == int(last_octet2):
                    ip = sheet.cell(row=x, column=y).value
                    mask = sheet.cell(row=x, column=y+2).value
                    ip_list.append(ip)

            # v - ip csg, v3- ip bs, v4 - mask
            elif "." in str(v3) and "255.255.255." in str(v4):
                last_octet2 = v3.split(".")[3]
                if int(last_octet1) + 1 == int(last_octet2):
                    ip = sheet.cell(row=x, column=y).value
                    mask = sheet.cell(row=x, column=y+2).value
                    ip_list.append(ip)            
            
        y += 1
        if y == 60:             # max limit
            y = 1
            x += 1
        if len(ip_list) == 6:   # max limit
            break 
        if x == 30:             # max limit
            print("----------------------------------------------------------------------")
            print("BREAK loop, check IP")
            print("----------------------------------------------------------------------")
            break

    conf_dic["ip"] = ip_list
    conf_dic["mask"] = mask


def check_ip(conf_dic):
    # goes before configuration function

    if len(conf_dic["ip"]) == 6:
        if "255.255.255" not in conf_dic["mask"]:
            conf_dic["errors"].append(f"check mask: {conf_dic['ip']}")

        last_octets = [i.split(".")[3] for i in conf_dic['ip']]

        if len(set(last_octets)) != 1:
            conf_dic["errors"].append(f"check last octets: {conf_dic['ip']}")

        third_octets = [int(i.split(".")[2]) for i in conf_dic['ip']]
        third_octets_diff = []        
        i = 0

        for i in range(5):
            d = third_octets[i+1] - third_octets[i]
            i += 1
            third_octets_diff.append(d)
        
        if len(set(third_octets_diff)) != 1:
            conf_dic["errors"].append(f"check octets diff: {conf_dic['ip']}")

    else:
        conf_dic["errors"].append(f"check len of IPs: {conf_dic['ip']}")


def read_template(conf_dic):
    template_output = ""
    env = Environment(loader=FileSystemLoader('./'))

    if conf_dic["ios_type"] == "cisco_ios":
        template = env.get_template("ios-template.txt")        
    elif conf_dic["ios_type"] == "cisco_xr":
        template = env.get_template("xr-template.txt")

    template_output = template.render(config = conf_dic)

    return template_output.splitlines()


def check_ip_duplication(usr, psw, conf_dic):
    # goes before configuration
    # check ip route in alma-agg-2
    ip = conf_dic['ip'][0]
    iplist = ip.split(".")

    net = ".".join([iplist[0], iplist[1], iplist[2], str(int(iplist[3])-1)])
    sh_ip = None

    try:
        connection = ConnectHandler(device_type="cisco_xr", ip="10.238.0.17", username=usr, password=psw)
        sh_ip = connection.send_command(f"show route vrf ALTEL_EPC_MA {net} {conf_dic['mask']}")
        connection.disconnect()
    except:
        conf_dic["errors"].append("alma-agg-2 connection error")

    if net in sh_ip:
        conf_dic["errors"].append(f"ip route duplication: {net}")


def configure(ssh_conn, commands, configuration_log, cfg):
    if cfg:

        if len(conf_dic["errors"]) > 0:
            print("----------------------------------------------------------------------\n"
                  "there are errors. conf is not loaded\n"
                  "----------------------------------------------------------------------")
        else:
            if len(commands) > 0:
                configuration_log.append(ssh_conn.send_config_set(commands))
                if conf_dic["ios_type"] == "cisco_ios":
                    try:
                        configuration_log.append(ssh_conn.save_config())
                    except Exception as err_msg:
                        configuration_log.append(f"COMMIT is OK after msg:{err_msg}")
                        configuration_log.append(ssh_conn.send_command("\n", expect_string=r"#"))
                
                elif conf_dic["ios_type"] == "cisco_xr":
                    configuration_log.append(ssh_conn.send_command("show configuration"))
                    configuration_log.append(ssh_conn.commit())
                    ssh_conn.exit_config_mode()

            else:
                print("----------------------------------------------------------------------\n"
                    "cfg is not needed")
    else:
        if len(commands) > 0:
            print("----------------------------------------------------------------------\n"
                  "candidate configuration:\n"
                  "----------------------------------------------------------------------")
            for line in commands:
                print(line)


def write_logs(cfg, commands):
    start_time = datetime.now()
    current_date = start_time.strftime("%Y.%m.%d")
    current_time = start_time.strftime("%H.%M")

    log_folder = Path(f"{Path.cwd()}/logs/{current_date}/")  # current dir / logs / date /
    log_folder.mkdir(exist_ok=True)
    
    config = log_folder / f"{current_time}_configuration_log.txt"
    config_file = open(config, "w")
    
    if cfg and commands:
        config_file.write("#" * 80 + "\n")
        config_file.write(f"########################################\n\n")
        config_file.write("".join(configuration_log))
        config_file.write("\n\n")

    config_file.close()
    if not cfg:
        config.unlink()         


def check_commit(configuration_log, commands, cfg):
    if cfg and len(conf_dic["errors"]) == 0:
        for i in configuration_log:
            if "%" in i:
                print("----------------------------------------------------------------------\n"
                    "# ERROR # CFG-COMMIT")

        for j in commands:
            if "!" not in j and "" != j and "no shutdown" not in j:
                if j not in "".join(configuration_log):
                    print("----------------------------------------------------------------------\n"
                        "# ERROR # Not all config is loaded. Check cfg log")


#######################################################################################
# ------------------------------              ----------------------------------------#
#######################################################################################

print("----------------------------------------------------------------------")
cfg = get_arguments(argv)
print("----------------------------------------------------------------------")
username, password = get_user_pw()
device = input("Enter device (ip or hostname): ")
set_ios(conf_dic, device)

try:
    ssh_conn, show_inf_desc = connect(username, password, conf_dic, device)
except:
    ssh_conn = False
    print("----------------------------------------------------------------------\n"
          "# ERROR # connection error\n"
          "----------------------------------------------------------------------")
    
if ssh_conn:
    region, hostname = define_hostname(ssh_conn, device)
    set_region(conf_dic, region, helper)
    print("----------------------------------------------------------------------")
    define_port_vlan(show_inf_desc, conf_dic)
    load_excel(conf_dic)
    check_ip(conf_dic)
    commands = read_template(conf_dic)
    check_ip_duplication(username, password, conf_dic)
    configure(ssh_conn, commands, configuration_log, cfg)
    ssh_conn.disconnect()
    write_logs(cfg, commands)
    check_commit(configuration_log, commands, cfg)

    if len(conf_dic["errors"]) > 0:
        print("----------------------------------------------------------------------\n"
              "ERRORS:\n")
        for e in conf_dic["errors"]:
            print(e)

        print("----------------------------------------------------------------------\n")

    else:
        print("----------------------------------------------------------------------\n"
              "success\n"
              "----------------------------------------------------------------------")

